from block import *
from consts.base_consts import *

filename = 'table.xlsx'

def process_blocks():
    global DOMEN_SELECT
    for domen in BASE_DOMEN.keys():
        print('process', domen)
        DOMEN_SELECT[0] = domen
        
        block2_main_1_2345() # ОКЕЙ
        block1_main_price_1234_234() # 28 ОКЕЙ
        block2_main_3_4() # ОКЕЙ
        block2_main_4_4() # ОКЕЙ
        block2_main_2_345() # ОКЕЙ
        block2_main_4_5() # ОКЕЙ
        block2_main_5_5() # ОКЕЙ
        block2_main_3_5() # ОКЕЙ
        block2_main_3_3() # ОКЕЙ
        block1_main_2345_2345() # ОКЕЙ
        
        
        block4_main_geo_12345_12345() # ОКЕЙ
        block4_main_geo_1_2345() # ОКЕЙ
        
        
        block4_geo_1234_1234() # ОКЕЙ
        block2_geo_main_1_2345() # ОКЕЙ

        block2_geo_main_4_4() # ОКЕЙ
        block2_geo_main_3_3() # ОКЕЙ
        block2_geo_main_2_3() # ОКЕЙ
        block2_geo_main_2_4() # ОКЕЙ

        block2_geo_main_4_5() # ОКЕЙ
        block2_geo_main_3_5() # ОКЕЙ
        block2_geo_main_2_5() # ОКЕЙ

        block1_geo_234_1234() # 24 ОКЕЙ
        block1_geo_1_234() # 25 пункт

        block1_main_234_1234() # 27 ОКЕЙ vacancy
        
        block5_main_12345_12345()
        
    block3_main_1234_1234() # 26 ОКЕЙ
    check_tag_url_block4() # ОКЕЙ
    with open('result.json', "w", encoding='UTF-8') as f:
        f.write(json.dumps(BASE_DATA, indent=4, ensure_ascii=False))

def check_tag_url_block4(name_block = 'block4'):
    for main_info in BASE_DATA:
        block4 = main_info['name_bloks'][name_block]["tag_url"]
        for block_info in block4:
            if main_info['level'] != block_info[3]:break
        else:
            for block_info in block4:
                block_info[0] = block_info[2]

def process_pagerank():
    global BASE_DATA
    if not len(BASE_DATA):
        with open(f'result.json', encoding='UTF-8') as f: BASE_DATA = json.load(f)
    import numpy as np
    from fast_pagerank import pagerank
    from scipy import sparse
    base_urls = tuple(map(lambda info: info['url'], BASE_DATA))

    page_counts = len(base_urls)
    
    result = list()
    
    for info_url in BASE_DATA:
        main_i = base_urls.index(info_url['url'])
        for i_block in range(1, 5):
            for tag_url in info_url['name_bloks'][f'block{i_block}']["tag_url"]:
                result.append([main_i, base_urls.index(tag_url[1])])
    if not len(result): return
    A = np.array(result)
    G = sparse.csr_matrix(([1] * len(A), (A[:,0], A[:,1])), shape=(page_counts, page_counts))  #shape=(4, 4))
    pr = pagerank(G, p=0.85)
    for info in zip(BASE_DATA, pr):
        info[0].update({'PR':info[1]})
    with open('result.json', "w", encoding='UTF-8') as f:
        f.write(json.dumps(BASE_DATA, indent=4, ensure_ascii=False))

def process_sheet(name_table = 'Data (все)'): #Вакансии мебель
    from openpyxl import load_workbook
    wb = load_workbook(filename, read_only= True)
    print(wb.sheetnames)
    sheet_data = wb[name_table] #Вакансии
    for row in tuple(sheet_data.rows)[1:]:
        add_info(row)

def convert_block(block):
    result = list()
    for tag in block['tag_url']:
        result.append(f'<a href="{tag[1]}">{tag[0]}</a>')
    if not len(result): return ''
    return f"""<h2 style="margin-bottom: 10px;">{block['name']}</h2><p>{'    ▪    '.join(result)}</p>"""

def get_base_text_cvs():
    with open('data_text.csv', encoding='UTF-8') as f:
        import re
        base = f.read().replace('\n', '')
        result = re.findall(r'"(\d{1,8})","(.*?)"/n', base)
        return dict(result)

def convert_html():
    global BASE_DATA
    if not len(BASE_DATA):
        with open(f'result.json', encoding='UTF-8') as f: BASE_DATA = json.load(f)
    base_text = get_base_text_cvs()
    sh = "update SeoData set seoStory = '{seoStory}'   where id = (select seoId from {mode_from} where id = {id});"
    name_froms = ('ContainerSection', 'Vacancy')
    base_result = list()
    for info in BASE_DATA:
        if info.get('is_price'): continue
        result = list()
        name_bloks = tuple(info["name_bloks"].values())
        for i_block in (1, 0, 2, 3, 4): # +4
            result.append(convert_block(name_bloks[i_block]))
        result = ''.join(result)
        if result == '': continue
        i_from = 0
        if info['url'].find('redsale.by/vacancy') != -1:# or result.find('redsale.by/repetitory') != -1: 
            i_from = 1
        result += base_text.get(str(info['id_container']), '')
        base_result.append(sh.format(id = info['id_container'], mode_from = name_froms[i_from], seoStory = result))
    with open('result.txt', 'w', encoding='UTF-8') as f:
        f.write('\n'.join(base_result))

def main():
    process_sheet()
    process_blocks()
    process_pagerank()
    convert_html()
    
    print('Ok!')

import json
import time

if __name__ == '__main__':
    time_start = time.time()
    main()
    print(time.time() - time_start)